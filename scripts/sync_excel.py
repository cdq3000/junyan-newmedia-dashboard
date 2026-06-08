from __future__ import annotations

import argparse
import json
import time
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


MONTHS = [f"{i}月" for i in range(1, 13)]
METRICS = {
    "liveSessions": "直播场次",
    "shortVideos": "短视频发布",
    "leads": "总线索量",
    "visits": "邀约到店",
    "orders": "新媒体订单",
    "orderShare": "新媒体订单占零售占比",
    "attritionRate": "门店离系率",
    "spend": "投流费用",
}

NAME_ALIASES = {
    "花都建设北零跑": "花都零跑",
    "清远港鸿": "清远港鸿零跑",
    "清远奇晟": "清远奇晟零跑",
    "清远英德": "清远英德零跑",
    "广花店": "广花零跑",
    "北站店": "北站零跑",
    "花都北站零跑": "北站零跑",
}

STORE_ORDER = [
    "花都广本",
    "清远广本",
    "汕头广本",
    "鑫海广本",
    "花都零跑",
    "白云零跑",
    "清远港鸿零跑",
    "清远奇晟零跑",
    "清远英德零跑",
    "河源零跑",
    "汕头零跑",
    "广花零跑",
    "北站零跑",
    "清远极氪",
    "清远智己",
    "江门大众",
    "花都传祺",
    "从化零跑",
]


def norm_store(value: Any) -> str | None:
    if value is None:
        return None
    name = str(value).strip()
    if not name:
        return None
    return NAME_ALIASES.get(name, name)


def num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, str) and value.startswith("#"):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def ensure_store(stores: dict[str, dict[str, Any]], name: str) -> dict[str, Any]:
    if name not in stores:
        stores[name] = {
            "name": name,
            "monthly": {
                month: {key: 0 for key in METRICS}
                for month in MONTHS
            },
        }
    return stores[name]


def read_row_months(ws, row: int, start_col: int) -> list[str]:
    months = []
    for offset in range(12):
        value = ws.cell(row=row, column=start_col + offset).value
        months.append(str(value).strip() if value else MONTHS[offset])
    return months


def read_metric_block(ws, stores: dict[str, dict[str, Any]], title_row: int, start_col: int, metric: str) -> None:
    months = read_row_months(ws, title_row + 1, start_col + 1)
    row = title_row + 2
    while row <= ws.max_row:
        name = norm_store(ws.cell(row=row, column=start_col).value)
        if not name:
            break
        store = ensure_store(stores, name)
        for offset, month in enumerate(months):
            value = num(ws.cell(row=row, column=start_col + 1 + offset).value)
            if value is not None and month in store["monthly"]:
                store["monthly"][month][metric] = value
        row += 1


def read_monthly_sheet(wb, stores: dict[str, dict[str, Any]]) -> None:
    ws = wb.worksheets[6]
    read_metric_block(ws, stores, 1, 1, "orders")
    read_metric_block(ws, stores, 1, 15, "orderShare")
    read_metric_block(ws, stores, 22, 1, "leads")
    read_metric_block(ws, stores, 22, 15, "visits")


def read_detail_sheet(wb, stores: dict[str, dict[str, Any]]) -> None:
    ws = wb.worksheets[7]
    current_store = None
    month_cols = []
    col = 5
    while col <= ws.max_column:
        raw = ws.cell(row=1, column=col).value
        label = str(raw).strip() if raw else ""
        if not label or label == "合计":
            break
        month_cols.append((label, col))
        col += 1
    if not month_cols:
        raise ValueError("No month columns found in detail sheet")
    metric_map = {
        "直播场次": "liveSessions",
        "短视频发布": "shortVideos",
        "投流费用": "spend",
    }
    for row in range(2, ws.max_row + 1):
        store_name = norm_store(ws.cell(row=row, column=2).value)
        if store_name:
            current_store = store_name
        metric_name = ws.cell(row=row, column=4).value
        metric = metric_map.get(str(metric_name).strip() if metric_name else "")
        if not current_store or not metric:
            continue
        store = ensure_store(stores, current_store)
        for month, col in month_cols:
            value = num(ws.cell(row=row, column=col).value)
            if value is not None:
                store["monthly"][month][metric] = value


def month_from_summary_title(value: Any) -> str | None:
    import re

    text = "" if value is None else str(value)
    matches = re.findall(r"(\d{1,2})月", text)
    return f"{int(matches[-1])}月" if matches else None


def read_latest_summary_sheet(wb, stores: dict[str, dict[str, Any]]) -> None:
    ws = wb.worksheets[2]
    month = month_from_summary_title(ws.cell(row=2, column=3).value) or month_from_summary_title(ws.cell(row=1, column=1).value)
    if not month:
        return
    mapping = {
        "leads": 8,
        "visits": 18,
        "orders": 19,
        "orderShare": 22,
    }
    for row in range(5, ws.max_row + 1):
        name = norm_store(ws.cell(row=row, column=2).value)
        if not name:
            continue
        if name == "合计":
            break
        store = ensure_store(stores, name)
        for metric, col in mapping.items():
            value = num(ws.cell(row=row, column=col).value)
            if value is not None:
                if metric == "orderShare":
                    store["monthly"][month][metric] = value
                else:
                    previous_total = sum(
                        (store["monthly"].get(item, {}) or {}).get(metric, 0) or 0
                        for item in MONTHS
                        if int(item.replace("月", "")) < int(month.replace("月", ""))
                    )
                    store["monthly"][month][metric] = max(0, value - previous_total)


def build_payload(excel_path: Path) -> dict[str, Any]:
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    stores: dict[str, dict[str, Any]] = {}
    read_monthly_sheet(wb, stores)
    read_detail_sheet(wb, stores)
    read_latest_summary_sheet(wb, stores)

    ordered = sorted(
        stores.values(),
        key=lambda item: (
            STORE_ORDER.index(item["name"]) if item["name"] in STORE_ORDER else 999,
            item["name"],
        ),
    )
    active_months = [
        month
        for month in MONTHS
        if any(any(store["monthly"][month].values()) for store in ordered)
    ]

    totals = {
        month: {key: 0 for key in METRICS}
        for month in active_months
    }
    for month in active_months:
        order_weight = 0
        weighted_share = 0
        for store in ordered:
            values = store["monthly"][month]
            for metric in totals[month]:
                if metric != "orderShare":
                    totals[month][metric] += values.get(metric, 0) or 0
            orders = values.get("orders", 0) or 0
            share = values.get("orderShare", 0) or 0
            order_weight += orders
            weighted_share += share * orders
        totals[month]["orderShare"] = weighted_share / order_weight if order_weight else 0

    latest_month = active_months[-1] if active_months else None
    previous_month = active_months[-2] if len(active_months) > 1 else None
    return {
        "title": "骏延集团新媒体数据",
        "sourceFile": str(excel_path),
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "months": active_months,
        "latestMonth": latest_month,
        "previousMonth": previous_month,
        "metrics": METRICS,
        "stores": ordered,
        "totals": totals,
    }


def write_payload(excel_path: Path, output_path: Path) -> None:
    payload = build_payload(excel_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[{datetime.now().strftime('%H:%M:%S')}] synced {excel_path.name} -> {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync Junyan Excel data to dashboard JSON.")
    parser.add_argument("--excel", default=str(Path.home() / "Desktop" / "2026年1-5月骏延集团新媒体数据表(总.xlsx"))
    parser.add_argument("--out", default=str(Path(__file__).resolve().parents[1] / "data" / "dashboard-data.json"))
    parser.add_argument("--watch", action="store_true")
    parser.add_argument("--interval", type=float, default=5)
    args = parser.parse_args()

    excel_path = Path(args.excel)
    output_path = Path(args.out)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    last_mtime = None
    while True:
        mtime = excel_path.stat().st_mtime
        if mtime != last_mtime:
            write_payload(excel_path, output_path)
            last_mtime = mtime
        if not args.watch:
            break
        time.sleep(args.interval)


if __name__ == "__main__":
    main()
